import requests
import json
import matplotlib.pyplot as plt
import socket
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
import requests
import statistics
import re

def direccion_a_coordenadas(direccion, user_agent="geolocalizador-app/1.0 (dev@miapp.com)"):
    """
    Convierte una dirección en coordenadas geográficas usando la API pública de Nominatim (OpenStreetMap).
    
    Parámetros:
    - direccion (str): Dirección a geocodificar.
    - user_agent (str): Identificador de la aplicación para cumplir con la política de uso de Nominatim.

    Retorna:
    - (latitud, longitud) como tupla de float si encuentra resultado.
    - (None, None) si no hay resultados o hay un error.

    Ejemplo:
    >>> lat, lon = direccion_a_coordenadas("Parque Fundidora, Monterrey, México")
    >>> print(lat, lon)
    """

    url = "https://nominatim.openstreetmap.org/search"
    params = {"q": direccion, "format": "json"}
    headers = {"User-Agent": user_agent}

    try:
        response = requests.get(url, params=params, headers=headers, timeout=10)
        response.raise_for_status()
        data = response.json()

        if data:
            return data
        else:
            return None, None

    except requests.RequestException as e:
        print(f"Error en la solicitud: {e}")
        return None, None


def obtener_datos_pvgis(lat, lon, potencia_kw, perdidas):
    url = "https://re.jrc.ec.europa.eu/api/v5_3/PVcalc"
    params = {
        "lat": lat,
        "lon": lon,
        "peakpower": potencia_kw,
        "loss": perdidas,
        "outputformat": "json"
    }
    response = requests.get(url, params)
    if response.status_code == 200:
        try:
            return response.json()
        except Exception as e:
            print("Error al procesar JSON:", e)
    else:
        print(f"Error en la API: código {response.status_code}")
    return None

def extraer_energia_mensual(data_json):
    meses_nombre = {
        1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
        5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
        9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
    }
    energia_mensual = {}
    meses = data_json['outputs']['monthly']['fixed']
    for diccionario in meses:
        mes = diccionario['month']
        energia = diccionario['E_m']
        nombre_mes = meses_nombre[mes]
        energia_mensual[nombre_mes] = energia
    return energia_mensual

def graficar_energia_mensual(energia_mensual):
    meses = list(energia_mensual.keys())
    energia = list(energia_mensual.values())
    
    plt.figure(figsize=(10, 6))
    plt.bar(meses, energia)
    plt.xlabel('Mes')
    plt.ylabel('Energía mensual (kWh)')
    plt.title('Producción mensual de energía')
    plt.xticks(rotation=45)
    plt.grid(axis='y', linestyle='--', alpha=0.5)
    plt.tight_layout()
    plt.show()

import requests

def elegir_direccion(data):
    print("Resultados encontrados:\n")
    for i in range(min(len(data), 10)):
        print(f"{i + 1}. {data[i]['display_name']}\n")

    while True:
        seleccion = input("Elige una opción (1-10) o escribe 'n' para intentar con otra dirección: ").strip().lower()
        if seleccion == 'n':
            return None, None, None
        try:
            seleccion = int(seleccion)
            if 1 <= seleccion <= min(len(data), 10):
                eleccion = data[seleccion - 1]
                lat = float(eleccion['lat'])
                lon = float(eleccion['lon'])
                return lat, lon, eleccion["display_name"]
            else:
                print("❗ Elige un número dentro del rango.")
        except ValueError:
            print("❗ Entrada no válida. Escribe un número o 'n'.")


def obtener_coordenadas_interactivas():
    


    while True:
        direccion = input("Ingresa una dirección (o escribe 'salir'): ").strip()
        if direccion.lower() == "salir":
            return exit()
        if not direccion:
            print("La dirección no puede estar vacía.")
            continue

        try:
            params = {"q": direccion, "format": "json", "limit": 10}
            headers = {"User-Agent": "geoapp/1.0 (dev@miapp.com)"}
            response = requests.get("https://nominatim.openstreetmap.org/search", params=params, headers=headers)
            response.raise_for_status()
            data = response.json()

            if not data:
                print("No se encontraron resultados. Intenta con una dirección más específica.")
                continue

            lat, lon, ubi = elegir_direccion(data)
            if lat and lon:
                return lat, lon, ubi
            else:
                print("Puedes ingresar otra dirección si quieres ser más específico.")

        except requests.RequestException:
            print("Hubo un error de conexión. Intenta más tarde.")
            return None, None
        
def cargar_json_local():
    """
    Carga y retorna el contenido del archivo JSON ubicado en 'datos/prueba.json'.
    Si el archivo no existe o es inválido, se muestra un mensaje de error.
    """
    

    try:
        with open("prueba.json", "r", encoding="utf-8") as archivo:
            return json.load(archivo)

    except FileNotFoundError:
        print("⚠️ El archivo 'prueba.json' no se encontró.")
        return None

    except json.JSONDecodeError:
        print("⚠️ El archivo no contiene un JSON válido.")
        return None




def hay_internet(host="8.8.8.8", port=53, timeout=3):
    """
    Verifica si hay conexión a Internet intentando abrir un socket TCP a un servidor DNS.
    Por defecto, usa Google DNS (8.8.8.8:53).

    Retorna:
        True si hay conexión, False si no.
    """
    try:
        socket.setdefaulttimeout(timeout)
        socket.socket(socket.AF_INET, socket.SOCK_STREAM).connect((host, port))
        return True
    except socket.error:
        return False


def exportar_a_excel(energia_mensual, ubicacion="", nombre_archivo="informe_solar_estadisticas.xlsx", precio_kwh=2.5):
    # Crear libro y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Energía mensual"

    # Información inicial
    ws["A1"] = "Ubicación:"
    ws["B1"] = ubicacion
    ws["A2"] = "Precio por kWh:"
    ws["B2"] = f"${precio_kwh:.2f} MXN"

    # Encabezados
    ws["A4"] = "Mes"
    ws["B4"] = "Energía (kWh)"
    ws["C4"] = "Ahorro estimado (MXN)"

    meses = list(energia_mensual.keys())
    valores_kwh = list(energia_mensual.values())
    valores_mxn = [e * precio_kwh for e in valores_kwh]

    # Insertar datos
    for i, mes in enumerate(meses):
        fila = i + 5
        ws[f"A{fila}"] = mes.capitalize()
        ws[f"B{fila}"] = valores_kwh[i]
        ws[f"C{fila}"] = valores_mxn[i]

    # Calcular estadísticas
    def safe_mode(lista):
        return statistics.mode(lista) if len(set(lista)) < len(lista) else "Sin moda"

    fila_est = len(meses) + 7
    ws[f"A{fila_est}"] = "Estadísticas de Energía:"
    ws[f"A{fila_est+1}"] = "Media"
    ws[f"B{fila_est+1}"] = round(statistics.mean(valores_kwh), 2)
    ws[f"A{fila_est+2}"] = "Mediana"
    ws[f"B{fila_est+2}"] = round(statistics.median(valores_kwh), 2)
    ws[f"A{fila_est+3}"] = "Moda"
    ws[f"B{fila_est+3}"] = safe_mode(valores_kwh)
    ws[f"A{fila_est+4}"] = "Desviación estándar"
    ws[f"B{fila_est+4}"] = round(statistics.stdev(valores_kwh), 2)

    fila_est += 6
    ws[f"A{fila_est}"] = "Estadísticas de Ahorro:"
    ws[f"A{fila_est+1}"] = "Media"
    ws[f"B{fila_est+1}"] = round(statistics.mean(valores_mxn), 2)
    ws[f"A{fila_est+2}"] = "Mediana"
    ws[f"B{fila_est+2}"] = round(statistics.median(valores_mxn), 2)
    ws[f"A{fila_est+3}"] = "Moda"
    ws[f"B{fila_est+3}"] = safe_mode(valores_mxn)
    ws[f"A{fila_est+4}"] = "Desviación estándar"
    ws[f"B{fila_est+4}"] = round(statistics.stdev(valores_mxn), 2)

    # Gráfica de energía
    img_energia = "grafica_energia.png"
    plt.figure(figsize=(10, 4))
    plt.bar(meses, valores_kwh, color="skyblue")
    plt.xlabel("Mes")
    plt.ylabel("Energía (kWh)")
    plt.title("Producción mensual de energía")
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.grid(axis='y', linestyle='--', alpha=0.5)
    plt.savefig(img_energia)
    plt.close()

    # Gráfica de ahorro
    img_ahorro = "grafica_ahorro.png"
    plt.figure(figsize=(10, 4))
    plt.bar(meses, valores_mxn, color="green")
    plt.xlabel("Mes")
    plt.ylabel("Ahorro estimado (MXN)")
    plt.title("Ahorro mensual estimado ($2.50 por kWh)")
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.grid(axis='y', linestyle='--', alpha=0.5)
    plt.savefig(img_ahorro)
    plt.close()

    # Insertar imágenes correctamente
    imagen1 = ExcelImage(img_energia)
    imagen1.anchor = "E4"
    ws.add_image(imagen1)

    imagen2 = ExcelImage(img_ahorro)
    imagen2.anchor = "E25"
    ws.add_image(imagen2)

    # Guardar archivo
    ruta_salida =  nombre_archivo
    wb.save(ruta_salida)
    return ruta_salida



def pedir_potencia():
    patron = r"^(?!0+(?:\.0+)?$)\d+(\.\d+)?$"  # evita exactamente 0 o 0.0

    while True:
        entrada = input("Ingrese la potencia del sistema en kW (mayor a 0): ").strip()

        if not re.match(patron, entrada):
            print("Potencia no válida. Debe ser un número mayor a 0.")
            continue

        try:
            potencia = float(entrada)
            if potencia <= 0:
                print("La potencia debe ser mayor que cero.")
            else:
                return potencia
        except ValueError:
            print("Entrada no válida. Intente de nuevo.")
