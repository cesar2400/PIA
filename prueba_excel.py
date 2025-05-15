import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
import matplotlib.pyplot as plt

def exportar_a_excel(energia_mensual, ubicacion="", nombre_archivo="informe_solar.xlsx"):
    """
    Exporta la energía mensual a un archivo Excel con una gráfica y ubicación al inicio.
    """

    # Crear libro de Excel y hoja activa
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Energía mensual"

    # Agregar ubicación en la parte superior
    ws["A1"] = "Ubicación:"
    ws["B1"] = ubicacion

    # Encabezados de la tabla
    ws["A3"] = "Mes"
    ws["B3"] = "Energía (kWh)"

    # Insertar los datos debajo
    fila = 4
    for mes, energia in energia_mensual.items():
        ws[f"A{fila}"] = mes.capitalize()
        ws[f"B{fila}"] = energia
        fila += 1

    # Crear gráfica y guardarla como imagen
    meses = list(energia_mensual.keys())
    valores = list(energia_mensual.values())

    plt.figure(figsize=(10, 6))
    plt.bar(meses, valores)
    plt.xlabel("Mes")
    plt.ylabel("Energía mensual (kWh)")
    plt.title("Producción mensual de energía")
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.grid(axis='y', linestyle='--', alpha=0.5)
    nombre_imagen = "grafica_solar.png"
    plt.savefig(nombre_imagen)
    plt.close()

    # Insertar imagen en Excel
    img = ExcelImage(nombre_imagen)
    img.anchor = f"D3"
    ws.add_image(img)

    # Guardar el archivo
    wb.save(nombre_archivo)
    print(f"✅ Informe exportado como '{nombre_archivo}'")

