import os
from openpyxl import Workbook as work
from openpyxl import load_workbook as lw
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
import matplotlib.pyplot as plt
from os.path import exists
import pandas as pd
import os 
import json
from re import search
import pyautogui
import subprocess
import psutil
import datetime
import time
import statistics


import main 

def extraer_fecha_hora():
	"""
	Devuelve la fecha y la hora del momento de su ejecución.
	"""
	ahora = datetime.datetime.now()
	fecha = ahora.strftime("%d/%m/%Y")
	hora = ahora.strftime("%H:%M:%S")
	
	return fecha,hora

def obtener_nombre_proceso(nombre_parcial_aplicacion):
	nombres_encontrados = []
	for proc in psutil.process_iter(['pid', 'name']):
		if nombre_parcial_aplicacion.lower() in proc.info['name'].lower():
			nombres_encontrados.append(proc.info['name'])
	return nombres_encontrados

def cerrar_proceso_windows(nombre_proceso):
	"""
	Intenta terminar un proceso por su nombre en Windows.
	Ejemplo: Si sabes que la aplicación se llama "mi_aplicacion.exe"
		cerrar_proceso_windows("mi_aplicacion.exe")
	"""
	try:
		subprocess.run(['taskkill', '/F', '/IM', f'{nombre_proceso}'], check=True)
	except subprocess.CalledProcessError as e:
		print(f"Error al intentar terminar el proceso '{nombre_proceso}': {e}")
	except FileNotFoundError:
		print("El comando 'taskkill' no se encontró (esto debería ser raro en Windows).")


def limpiar_pantalla():
	"""
	Esta función solo sirve para el sistema operativo de Windows cmd.
	Facilita la visualización de los datos pero, a no ser que se tengan almacenados los datos de antemano, se perderan.

	"""
	os.system('cls' if os.name == 'nt' else 'clear')
	
def pedir_numero_int(numero = None, mensaje = "Numero: "):
	"""
	Esta función se emplea en situaliones en las que es necesario obtener un numero entero.
	"""
	if numero:
		num_match = search(r"\d+",numero)
	else:
		num_match = None
	while not num_match or num_match.group() != numero:
		numero = input(mensaje)
		num_match = search(r"\d+",numero)
		#print(numero,num_match)
		
		if not num_match == None and num_match.group() == numero:
			break
	return int(numero)
	
def abrir_archivo(ruta_archivo,nombre_parcial  = None): 
	"""
	Literalmente una función que abre y luego cierra automaticamente un archivo desde el sistema operativo de windows.
	Declaración de variables:
		ruta_archivo: Reemplaza con la ruta real de tu archivo con el formato r'*Ruta_especifica*' 
	"""
	print("Presione enter si desea poner un cronometro especifico, \"5\" si desea que solo espere los 5 minutos preestablesidos o cualquier otra tecla no realizar el proceso.")
	sn = input(">>> ")
	if sn == "":
		n = int(input("Introduzca el tiempo de espera en minutos. \n>>> "))
	elif sn == "5":
		n = None
	else:
		n = "p"
	try:
		os.startfile(ruta_archivo)
		print(f"Abriendo el archivo: {ruta_archivo}")
	except FileNotFoundError:
		print(f"El archivo no fue encontrado: {ruta_archivo}")
	except OSError as e:
		print(f"No se pudo abrir el archivo. Error: {e}")
	
	if nombre_parcial:
		procesos = obtener_nombre_proceso(nombre_parcial)
	elif n != "":
		print("No se introdujo el nombre de la aplicación.")
		return
	
	if not n and procesos:
		pausar()
		for i in range(len(procesos)):
			cerrar_proceso_windows(procesos[i])
		print(f"Archivo {ruta_archivo} cerrado")
	elif n != "p" and procesos:
		pausar(n)
		for i in range(len(procesos)):
			cerrar_proceso_windows(procesos[i])
		print(f"Archivo {ruta_archivo} cerrado")
	

def pausar(n = None):
	if not n:
		print("Tiene 5 minutos.")
		for i in range(5):
			time.sleep(60)
	elif n:
		for i in range(n):
			time.sleep(60)
		

def abrir_crear(nombre_archivo = None,nombre_hoja = None,nueva_hoja = None):
	if nombre_archivo and exists(nombre_archivo):
		try:
			archivo = lw(nombre_archivo)
		except Exception as e:
			archivo = work()
	else:
		archivo = work()
		
	if not (nombre_hoja and nueva_hoja):
		hoja = archivo.active
	elif nueva_hoja and nombre_hoja:
		archivo.create_sheet(nombre_hoja)
		hoja = archivo[nombre_hoja]
	elif nueva_hoja:
		archivo.create_sheet('Sheet')
		hoja = archivo['Sheet']
	elif nombre_hoja and nombre_hoja in archivo:
		hoja = archivo[nombre_hoja]
	elif nombre_hoja:
		hoja = archivo.active
		hoja.title = nombre_hoja
	else:
		hoja = archivo.active
	return archivo,hoja


def contenido_excel(nombre_archivo = None,nombre_hoja = None):
	archivo,hoja = abrir_crear(nombre_archivo,nombre_hoja)
	list_name = archivo.sheetnames
	archivo.save(nombre_archivo)

	datos = []
	for i in range(len(list_name)):
		datos.append(pd.read_excel(nombre_archivo,list_name[i]))
		datos[i] = datos[i].to_dict(orient='list')
		
	return datos	
	
def guardar_excel(nombre_archivo = None,nombre_hoja = None, datos = None):
	archivo,hoja = abrir_crear(nombre_archivo,nombre_hoja,nueva_hoja = True)
	list_name = archivo.sheetnames
	archivo.save(nombre_archivo)
	
	if type(datos) == list:
		aux = []
		for i in range(len(datos)):
			aux.append(pd.DataFrame(datos[i]))
		guardar_varios(nombre_archivo,list_name,aux)
	elif type(datos) == dict:
		df = pd.DataFrame(datos)
		with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
			df.to_excel(writer, sheet_name=nombre_hoja, index=False)
	
def guardar_varios(nombre_archivo,nombres_hojas,dataframes):
	if len(dataframes) != len(nombres_hojas):
		raise ValueError("La lista de DataFrames y la lista de nombres de hojas deben tener la misma longitud.")
	try:
		# Crear un objeto ExcelWriter para gestionar la escritura en el archivo
		with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
			# Iterar sobre los DataFrames y sus nombres de hoja correspondientes
			for i, df in enumerate(dataframes):
				nombre_hoja = nombres_hojas[i]
				df.to_excel(writer, sheet_name=nombre_hoja, index=False)  # Escribir el DataFrame en la hoja
			print(f"Se han guardado los DataFrames en el archivo '{nombre_archivo}'.")
		
	except Exception as e:
		print(f"Ocurrió un error al guardar los DataFrames en Excel: {e}")
def guardar_estadisticas(energia_mensual, ubicacion="",fecha_hora: tuple = None, nombre_archivo= None, nombre_hoja = "Caso estándar"):
	"""
	Exporta la energía mensual a un archivo Excel con una gráfica y ubicación al inicio.
	
	"""
	# Crear libro de Excel y hoja activa
	wb,ws = abrir_crear(nombre_archivo,nombre_hoja)
	

	# Agregar ubicación en la parte superior
	ws["A1"] = "Ubicación:"
	ws["B1"] = ubicacion
	ws["D1"] = "Fecha:"
	ws["E1"] = fecha_hora[0]
	ws["G1"] = "Hora:"
	ws["H1"] = fecha_hora[1]
	
	# Encabezados de la tabla
	ws["A3"] = "Mes"
	ws["B3"] = "Energía (kWh)"
	ws["C3"] = "Valor monetario"

	# Insertar los datos debajo
	fila = 4
	energias = []
	for mes, energia in energia_mensual.items():
		ws[f"A{fila}"] = mes.capitalize()
		ws[f"B{fila}"] = energia
		energias.append(energia)
		ws[f"C{fila}"] = energia * 2.5
		fila += 1
	
	ws[f"A{fila + 1}"] = "Datos estadisticos."
	ws[f"B{fila + 1}"] = "Valor estadistico."
	ws[f"A{fila + 2}"] = "Mediana:"
	ws[f"B{fila + 2}"] = statistics.median(energias)
	
	ws[f"A{fila + 3}"] = "Media:" 
	ws[f"B{fila + 3}"] = statistics.mean(energias)
	
	ws[f"A{fila + 4}"] = "Desviación estándar:" 
	ws[f"B{fila + 4}"] = statistics.stdev(energias)
	
	# Crear gráfica y guardarla como imagen
	meses = list(energia_mensual.keys())
	valores = list(energia_mensual.values())

	plt.figure(figsize=(10, 6))
	plt.bar(meses, valores)
	plt.xlabel("Mes")
	plt.ylabel("Energía mensual (kWh)")
	plt.title("Producción mensual de energía")
	plt.xticks(rotation=45)
	plt.tight_layout()
	plt.grid(axis='y', linestyle='--', alpha=0.5)
	nombre_imagen = "grafica_solar.png"
	plt.savefig(nombre_imagen)
	plt.close()
	# Insertar imagen en Excel
	img = ExcelImage(nombre_imagen)
	img.anchor = f"E3"
	ws.add_image(img)
	
	# Guardar el archivo
	wb.save(nombre_archivo)
	print(f"Informe exportado como '{nombre_archivo}'")

def renovar_contenido(archivo_entrada, archivo_salida = None, nombre_hoja = None,nueva_hoja = True):
	"""
	Extrae el contenido de un archivo Excel y lo guarda en un nuevo archivo o en el mismo pero con otra hoja,
	manteniendo la estructura básica (ubicación, fecha, hora, tabla, estadísticas, gráfica).
	"""
	if not archivo_salida:
		archivo_salida = archivo_entrada
		nueva_hoja = True
	try:
		# 1. Cargar el archivo de entrada
		libro_entrada = openpyxl.load_workbook(archivo_entrada)
		hoja_entrada = libro_entrada.active

		# 2. Crear un nuevo libro de Excel y hoja activa
		libro_salida, hoja_salida = abrir_crear(archivo_salida,nombre_hoja,nueva_hoja)

		# --- 3. Copiar la información de la parte superior ---
		hoja_salida["A1"] = hoja_entrada["A1"].value
		hoja_salida["B1"] = hoja_entrada["B1"].value
		hoja_salida["D1"] = hoja_entrada["D1"].value
		hoja_salida["E1"] = hoja_entrada["E1"].value
		hoja_salida["G1"] = hoja_entrada["G1"].value
		hoja_salida["H1"] = hoja_entrada["H1"].value

		# --- 4. Copiar los encabezados de la tabla ---
		hoja_salida["A3"] = hoja_entrada["A3"].value
		hoja_salida["B3"] = hoja_entrada["B3"].value
		hoja_salida["C3"] = hoja_entrada["C3"].value

		# --- 5. Copiar los datos de la tabla de energía mensual ---
		fila_entrada = 4
		fila_salida = 4
		energia_mensual_extraida = {}
		while hoja_entrada[f"A{fila_entrada}"].value is not None:
			hoja_salida[f"A{fila_salida}"] = hoja_entrada[f"A{fila_entrada}"].value
			hoja_salida[f"B{fila_salida}"] = hoja_entrada[f"B{fila_entrada}"].value
			hoja_salida[f"C{fila_salida}"] = hoja_entrada[f"C{fila_entrada}"].value
			energia_mensual_extraida[hoja_entrada[f"A{fila_entrada}"].value.lower()] = hoja_entrada[f"B{fila_entrada}"].value
			fila_entrada += 1
			fila_salida += 1

		# --- 6. Copiar los datos estadísticos ---
		fila_estadisticos_entrada = fila_entrada + 1
		hoja_salida[f"A{fila_salida + 1}"] = hoja_entrada[f"A{fila_estadisticos_entrada}"].value
		hoja_salida[f"B{fila_salida + 1}"] = hoja_entrada[f"B{fila_estadisticos_entrada}"].value
		hoja_salida[f"A{fila_salida + 2}"] = hoja_entrada[f"A{fila_estadisticos_entrada + 1}"].value
		hoja_salida[f"B{fila_salida + 2}"] = hoja_entrada[f"B{fila_estadisticos_entrada + 1}"].value
		hoja_salida[f"A{fila_salida + 3}"] = hoja_entrada[f"A{fila_estadisticos_entrada + 2}"].value
		hoja_salida[f"B{fila_salida + 3}"] = hoja_entrada[f"B{fila_estadisticos_entrada + 2}"].value
		hoja_salida[f"A{fila_salida + 4}"] = hoja_entrada[f"A{fila_estadisticos_entrada + 3}"].value
		hoja_salida[f"B{fila_salida + 4}"] = hoja_entrada[f"B{fila_estadisticos_entrada + 3}"].value

		# --- 7. Re-crear y copiar la gráfica ---
		meses = list(energia_mensual_extraida.keys())
		valores = list(energia_mensual_extraida.values())

		plt.figure(figsize=(10, 6))
		plt.bar(meses, valores)
		plt.xlabel("Mes")
		plt.ylabel("Energía mensual (kWh)")
		plt.title("Producción mensual de energía")
		plt.xticks(rotation=45)
		plt.tight_layout()
		plt.grid(axis='y', linestyle='--', alpha=0.5)
		nombre_imagen_nuevo = "grafica_temporal.png" # Nombre temporal para la nueva gráfica
		plt.savefig(nombre_imagen_nuevo)
		plt.close()

		img_nueva = ExcelImage(nombre_imagen_nuevo)
		img_nueva.anchor = f"D3"
		hoja_salida.add_image(img_nueva)

		# --- 8. Guardar el nuevo archivo ---
		libro_salida.save(archivo_salida)

	except FileNotFoundError:
		print(f"Error: El archivo '{archivo_entrada}' no fue encontrado.")
	except Exception as e:
		print(f"Ocurrió un error: {e}")

def extraer_datos(archivo_json):
	with open(archivo_json,'r') as archivo:
		resultado = json.load(archivo)
	
	limpiar_pantalla()
	outputs = resultado['outputs']
	aux = outputs['monthly']['fixed']
	mes = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
	datos = []
	
	for i in range(len(aux)):
		datos.append(dict())
		datos[i]['Mes'] = mes[i] 
		for j in aux[i].keys():
			if j != "month":
				datos[i][j] = aux[i][j]
	return datos
	
def excel_main(energia_mensual,ubi,nombre_archivo = 'excel_usuario.xlsx'):
	limpiar_pantalla()
	
	print("Introduzca su nombre:")
	nombre_hoja = input(">>> ")
	
	aux = "archivo_aux.xlsx"

	procesos = obtener_nombre_proceso("Excel")
	
	for i in range(len(procesos)):
		cerrar_proceso_windows(procesos[i])
	time.sleep(2)
	# Esta parte fue añadida para evitar que el proceso se interumpa por que el archivo esté abierto.
	
	fecha_hora = extraer_fecha_hora()
	
	guardar_estadisticas(energia_mensual,ubi,fecha_hora, aux, nombre_hoja)
	
	renovar_contenido(aux,nombre_archivo,nombre_hoja)
	
	ruta_archivo = r'excel_usuario.xlsx' 
	abrir_archivo(ruta_archivo,"Excel")

if __name__ == "__main__":
	info = main.main()
	excel_main(info[1],info[0])

